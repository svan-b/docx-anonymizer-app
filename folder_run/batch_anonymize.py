#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Folder Batch Anonymization System
==================================
Processes entire folder structures using proven anonymization backend.
Safe, production-grade batch processing with per-folder image removal prompting.

Author: Derived from proven Streamlit app backend
Date: 2025-11-15
"""

import sys
import os
from pathlib import Path

# CRITICAL FIX: Apply OOXML int() conversion patches BEFORE importing Document/Presentation
# Fixes: ValueError: invalid literal for int() with base 10: '19.5'
# Must be imported before process_adobe_word_files and process_powerpoint
# See: ../fix_ooxml_int_conversion.py for details
sys.path.insert(0, str(Path(__file__).parent.parent))
from fix_ooxml_int_conversion import apply_ooxml_patches
apply_ooxml_patches()
import argparse
import logging
from pathlib import Path
from datetime import datetime
import subprocess
import shutil
from typing import Dict, List, Tuple, Optional
from collections import defaultdict
import time

# Add parent directory to path to import proven modules
sys.path.insert(0, str(Path(__file__).parent.parent))

# Import proven, battle-tested backend modules (NO MODIFICATIONS)
from process_adobe_word_files import (
    load_aliases_from_excel,
    categorize_and_sort_aliases,
    precompile_patterns,
    process_single_docx
)
from process_powerpoint import process_single_pptx
from process_excel import process_single_xlsx, process_single_xls

# ANSI color codes for terminal output
class Colors:
    HEADER = '\033[95m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'


class BatchStats:
    """Tracks comprehensive statistics during batch processing"""

    def __init__(self):
        self.start_time = datetime.now()
        self.files_processed = 0
        self.files_succeeded = 0
        self.files_failed = 0
        self.files_skipped = 0
        self.total_replacements = 0
        self.total_images_removed = 0
        self.pdf_successes = 0
        self.pdf_failures = 0

        # Detailed tracking
        self.file_details = []  # List of dicts for per-file stats
        self.folder_stats = defaultdict(lambda: {
            'files': 0, 'succeeded': 0, 'failed': 0, 'skipped': 0,
            'replacements': 0, 'images_removed': 0
        })
        self.replacement_frequency = defaultdict(int)  # Track which replacements used
        self.error_log = []  # Detailed error information
        self.copied_files = []  # Track non-processable files copied as-is
        self.file_replacement_details = []  # Per-file detailed replacements (v2.1)

    def add_file_result(self, file_path: Path, relative_path: Path, status: str,
                       replacements: int = 0, images_removed: int = 0,
                       processing_time: float = 0, error_msg: str = "",
                       replacement_details: dict = None):
        """Record result for a single file"""
        folder_name = str(relative_path.parent) if relative_path.parent != Path('.') else 'root'

        # Update counters
        self.files_processed += 1
        if status == 'success':
            self.files_succeeded += 1
            self.total_replacements += replacements
            self.total_images_removed += images_removed
        elif status == 'failed':
            self.files_failed += 1
        elif status == 'skipped':
            self.files_skipped += 1

        # Store detailed record
        self.file_details.append({
            'file_path': str(relative_path),
            'folder': folder_name,
            'filename': file_path.name,
            'extension': file_path.suffix,
            'status': status,
            'replacements': replacements,
            'images_removed': images_removed,
            'processing_time': processing_time,
            'error': error_msg
        })

        # Store per-file replacement details (v2.1)
        if replacement_details and status == 'success':
            self.file_replacement_details.append({
                'file_path': str(relative_path),
                'directory': folder_name,
                'filename': file_path.name,
                'details': replacement_details  # {original: count, ...}
            })

        # Update folder stats
        self.folder_stats[folder_name]['files'] += 1
        if status == 'success':
            self.folder_stats[folder_name]['succeeded'] += 1
            self.folder_stats[folder_name]['replacements'] += replacements
            self.folder_stats[folder_name]['images_removed'] += images_removed
        elif status == 'failed':
            self.folder_stats[folder_name]['failed'] += 1
        elif status == 'skipped':
            self.folder_stats[folder_name]['skipped'] += 1

        # Store error if present
        if error_msg:
            self.error_log.append({
                'file_path': str(relative_path),
                'error': error_msg,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            })

    def add_pdf_result(self, success: bool):
        """Track PDF conversion result"""
        if success:
            self.pdf_successes += 1
        else:
            self.pdf_failures += 1

    def get_elapsed_time(self) -> str:
        """Get formatted elapsed time"""
        elapsed = datetime.now() - self.start_time
        minutes, seconds = divmod(elapsed.total_seconds(), 60)
        return f"{int(minutes)}m {int(seconds)}s"

    def get_summary(self, include_pdf: bool = True) -> str:
        """Get formatted summary string"""
        success_rate = (self.files_succeeded / max(self.files_processed, 1)) * 100

        # Build PDF conversion section conditionally
        pdf_section = ""
        if include_pdf:
            pdf_section = f"""
{Colors.BOLD}PDF Conversion:{Colors.ENDC}
  {Colors.GREEN}✓ Succeeded:{Colors.ENDC}       {self.pdf_successes}
  {Colors.RED}✗ Failed:{Colors.ENDC}          {self.pdf_failures}
"""

        return f"""
{Colors.BOLD}{Colors.CYAN}╔════════════════════════════════════════════════════════════╗
║              BATCH PROCESSING SUMMARY                      ║
╚════════════════════════════════════════════════════════════╝{Colors.ENDC}

{Colors.BOLD}Files Processed:{Colors.ENDC}     {self.files_processed}
  {Colors.GREEN}✓ Succeeded:{Colors.ENDC}       {self.files_succeeded}
  {Colors.RED}✗ Failed:{Colors.ENDC}          {self.files_failed}
  {Colors.YELLOW}⊘ Skipped:{Colors.ENDC}         {self.files_skipped}

{Colors.BOLD}Anonymization:{Colors.ENDC}
  Replacements:       {self.total_replacements:,}
  Images Removed:     {self.total_images_removed:,}{pdf_section}
{Colors.BOLD}Performance:{Colors.ENDC}
  Success Rate:       {success_rate:.1f}%
  Total Time:         {self.get_elapsed_time()}
"""


class ProgressDisplay:
    """Real-time terminal progress display"""

    def __init__(self, total_files: int):
        self.total_files = total_files
        self.current = 0
        self.last_update = time.time()

    def update(self, current_file: str, stats: BatchStats):
        """Update progress display"""
        self.current += 1

        # Calculate progress
        progress = self.current / max(self.total_files, 1)
        bar_length = 40
        filled = int(bar_length * progress)
        bar = '█' * filled + '░' * (bar_length - filled)

        # Clear previous lines (move cursor up 5 lines and clear)
        if self.current > 1:
            sys.stdout.write('\033[5A\033[J')

        # Print updated display
        print(f"\n{Colors.BOLD}{Colors.BLUE}Progress: [{bar}] {progress*100:.1f}%{Colors.ENDC}")
        print(f"{Colors.BOLD}Files:{Colors.ENDC} {self.current}/{self.total_files} | "
              f"{Colors.GREEN}Success: {stats.files_succeeded}{Colors.ENDC} | "
              f"{Colors.RED}Failed: {stats.files_failed}{Colors.ENDC} | "
              f"{Colors.YELLOW}Skipped: {stats.files_skipped}{Colors.ENDC}")
        print(f"{Colors.BOLD}Stats:{Colors.ENDC} Replacements: {stats.total_replacements:,} | "
              f"Images Removed: {stats.total_images_removed:,}")
        print(f"{Colors.BOLD}Time:{Colors.ENDC} {stats.get_elapsed_time()}")
        print(f"{Colors.CYAN}Current:{Colors.ENDC} {current_file[:70]}")

        sys.stdout.flush()


def setup_logging(log_dir: Path) -> logging.Logger:
    """Setup dual logging (file + console)"""
    log_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    log_file = log_dir / f"batch_run_{timestamp}.log"

    # Create logger
    logger = logging.getLogger('batch_anonymizer')
    logger.setLevel(logging.DEBUG)

    # File handler (detailed)
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_formatter = logging.Formatter(
        '%(asctime)s | %(levelname)-8s | %(message)s',
        datefmt='%Y-%m-%d %H:%M:%S'
    )
    file_handler.setFormatter(file_formatter)

    # Console handler (errors only)
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.ERROR)
    console_formatter = logging.Formatter('%(levelname)s: %(message)s')
    console_handler.setFormatter(console_formatter)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    logger.info(f"{'='*60}")
    logger.info(f"Batch Anonymization Session Started")
    logger.info(f"Log file: {log_file}")
    logger.info(f"{'='*60}")

    return logger


def get_folder_info(folder_path: Path, input_dir: Path) -> Dict:
    """Get information about a folder for prompting"""
    # Find all processable files in this folder tree (recursive)
    docx_files = list(folder_path.rglob('*.docx'))
    xlsx_files = list(folder_path.rglob('*.xlsx'))
    xlsm_files = list(folder_path.rglob('*.xlsm'))  # Excel macro-enabled workbooks
    pptx_files = list(folder_path.rglob('*.pptx'))
    doc_files = list(folder_path.rglob('*.doc'))
    xls_files = list(folder_path.rglob('*.xls'))
    ppt_files = list(folder_path.rglob('*.ppt'))

    # Also find non-processable files
    pdf_files = list(folder_path.rglob('*.pdf'))
    png_files = list(folder_path.rglob('*.png'))
    jpg_files = list(folder_path.rglob('*.jpg'))
    jpeg_files = list(folder_path.rglob('*.jpeg'))

    # Filter out tracker files from processable files
    processable_files = docx_files + xlsx_files + xlsm_files + pptx_files + doc_files + xls_files + ppt_files
    processable_files = [f for f in processable_files if 'tracker' not in f.name.lower() and 'anon tracker' not in f.name.lower()]

    non_processable_files = pdf_files + png_files + jpg_files + jpeg_files

    # Get file type breakdown
    type_counts = {
        'DOCX': len(docx_files) + len(doc_files),
        'XLSX': len(xlsx_files) + len(xls_files) + len(xlsm_files),
        'PPTX': len(pptx_files) + len(ppt_files)
    }

    # Get non-processable type breakdown
    non_processable_counts = {
        'PDF': len(pdf_files),
        'PNG': len(png_files),
        'JPG/JPEG': len(jpg_files) + len(jpeg_files)
    }

    # Filter to only non-zero counts
    non_processable_counts = {k: v for k, v in non_processable_counts.items() if v > 0}

    # Get sample filenames (first 5 processable)
    sample_files = [f.name for f in processable_files[:5]]

    # Get subdirectories in this folder
    subdirs = [d.relative_to(folder_path) for d in folder_path.rglob('*') if d.is_dir()]

    # Estimate processing time (rough: 5 seconds per file)
    est_time_seconds = len(processable_files) * 5
    est_minutes = est_time_seconds / 60

    # Get relative path
    try:
        relative_path = folder_path.relative_to(input_dir)
    except ValueError:
        relative_path = folder_path

    return {
        'path': relative_path,
        'file_count': len(processable_files),
        'type_counts': type_counts,
        'non_processable_counts': non_processable_counts,
        'non_processable_total': len(non_processable_files),
        'sample_files': sample_files,
        'subdirs_count': len(subdirs),
        'est_minutes': est_minutes,
        'has_warnings': len(non_processable_files) > 0
    }


def prompt_for_image_removal(folder_info: Dict, auto_mode: Optional[bool] = None) -> Tuple[bool, Optional[bool]]:
    """
    Prompt user whether to remove images from this folder

    Returns:
        (remove_images, auto_mode)
        auto_mode = True: auto-yes for all remaining
        auto_mode = False: auto-no for all remaining
        auto_mode = None: continue prompting
    """
    if auto_mode is not None:
        return auto_mode, auto_mode

    print(f"\n{Colors.BOLD}{Colors.CYAN}{'─'*60}{Colors.ENDC}")
    print(f"{Colors.BOLD}Folder:{Colors.ENDC} {folder_info['path']}")

    # Show subdirectory count
    if folder_info['subdirs_count'] > 0:
        print(f"{Colors.BOLD}Subdirectories:{Colors.ENDC} {folder_info['subdirs_count']}")

    print(f"{Colors.BOLD}Processable Files:{Colors.ENDC} {folder_info['file_count']} total")

    # Show file type breakdown
    type_str = " | ".join([f"{k}: {v}" for k, v in folder_info['type_counts'].items() if v > 0])
    if type_str:
        print(f"{Colors.BOLD}Types:{Colors.ENDC} {type_str}")

    # INFO: Show non-processable files if present
    if folder_info['has_warnings']:
        print(f"\n{Colors.CYAN}{Colors.BOLD}ℹ INFO: This folder contains non-processable files:{Colors.ENDC}")
        for file_type, count in folder_info['non_processable_counts'].items():
            print(f"  {Colors.CYAN}• {file_type}: {count} file(s) (will be COPIED as-is){Colors.ENDC}")
        print(f"{Colors.CYAN}  Total non-processable: {folder_info['non_processable_total']}{Colors.ENDC}")

    # Show sample filenames
    if folder_info['sample_files']:
        print(f"\n{Colors.BOLD}Sample processable files:{Colors.ENDC}")
        for fname in folder_info['sample_files']:
            print(f"  • {fname}")

    # Show estimated time
    print(f"\n{Colors.BOLD}Est. time:{Colors.ENDC} ~{folder_info['est_minutes']:.1f} minutes")
    print(f"{Colors.CYAN}{'─'*60}{Colors.ENDC}")

    # Prompt
    while True:
        response = input(f"\n{Colors.BOLD}Remove images from this folder?{Colors.ENDC} "
                        f"[{Colors.GREEN}y{Colors.ENDC}es / "
                        f"{Colors.RED}n{Colors.ENDC}o / "
                        f"{Colors.CYAN}a{Colors.ENDC}uto-yes / "
                        f"{Colors.YELLOW}s{Colors.ENDC}kip / "
                        f"{Colors.RED}q{Colors.ENDC}uit]: ").lower().strip()

        if response in ['y', 'yes']:
            return True, None
        elif response in ['n', 'no']:
            return False, None
        elif response in ['a', 'auto', 'auto-yes']:
            print(f"{Colors.CYAN}→ Auto-yes mode activated for all remaining folders{Colors.ENDC}")
            return True, True
        elif response in ['s', 'skip']:
            print(f"{Colors.YELLOW}→ Skipping this folder{Colors.ENDC}")
            return None, None  # None means skip
        elif response in ['q', 'quit']:
            print(f"{Colors.RED}→ Quitting batch processing{Colors.ENDC}")
            sys.exit(0)
        else:
            print(f"{Colors.RED}Invalid response. Please enter y/n/a/s/q{Colors.ENDC}")


def convert_legacy_format(file_path: Path, output_dir: Path, logger: logging.Logger) -> Optional[Path]:
    """
    Convert legacy formats (.doc, .xls, .ppt) to modern formats using LibreOffice

    Returns:
        Path to converted file, or None if conversion failed
    """
    extension = file_path.suffix.lower()

    # Create temp directory for conversion
    temp_dir = output_dir / '.temp_conversions'
    temp_dir.mkdir(parents=True, exist_ok=True)

    # Determine output format
    format_map = {
        '.doc': 'docx',
        '.xls': 'xlsx',
        '.ppt': 'pptx'
    }

    if extension not in format_map:
        return file_path  # Not a legacy format

    output_format = format_map[extension]
    output_file = temp_dir / f"{file_path.stem}.{output_format}"

    try:
        logger.info(f"Converting {file_path.name} from {extension} to .{output_format}")

        # LibreOffice conversion command
        cmd = [
            'libreoffice',
            '--headless',
            '--convert-to', output_format,
            '--outdir', str(temp_dir),
            str(file_path)
        ]

        # Increase timeout for Excel files (can be large/complex)
        timeout_seconds = 600 if extension == '.xls' else 300

        result = subprocess.run(
            cmd,
            timeout=timeout_seconds,
            capture_output=True,
            text=True
        )

        if result.returncode == 0 and output_file.exists():
            logger.info(f"Successfully converted to {output_file.name}")
            return output_file
        else:
            # Log detailed error information
            error_details = f"Return code: {result.returncode}"
            if result.stderr:
                error_details += f"\nStderr: {result.stderr.strip()}"
            if result.stdout:
                error_details += f"\nStdout: {result.stdout.strip()}"
            logger.error(f"Conversion failed for {file_path.name}: {error_details}")

            # For .xls files, warn that they cannot be processed
            if extension == '.xls':
                logger.warning(f"Legacy .xls file cannot be converted. File will be copied as-is to output (not anonymized): {file_path.name}")

            return None

    except subprocess.TimeoutExpired:
        logger.error(f"Conversion timeout ({timeout_seconds}s) for {file_path.name}")
        return None
    except Exception as e:
        logger.error(f"Conversion error for {file_path.name}: {str(e)}")
        return None


def convert_to_pdf(file_path: Path, pdf_output_dir: Path, logger: logging.Logger) -> bool:
    """
    Convert document to PDF using LibreOffice

    Returns:
        True if conversion succeeded, False otherwise
    """
    try:
        # Ensure output directory exists
        pdf_output_dir.mkdir(parents=True, exist_ok=True)

        # Expected PDF filename
        pdf_file = pdf_output_dir / f"{file_path.stem}.pdf"

        # LibreOffice PDF conversion
        cmd = [
            'libreoffice',
            '--headless',
            '--convert-to', 'pdf',
            '--outdir', str(pdf_output_dir),
            str(file_path)
        ]

        result = subprocess.run(
            cmd,
            timeout=300,  # 5 minute timeout
            capture_output=True,
            text=True
        )

        if result.returncode == 0 and pdf_file.exists():
            logger.debug(f"PDF conversion successful: {pdf_file.name}")
            return True
        else:
            logger.warning(f"PDF conversion failed for {file_path.name}: {result.stderr}")
            return False

    except subprocess.TimeoutExpired:
        logger.warning(f"PDF conversion timeout for {file_path.name}")
        return False
    except Exception as e:
        logger.warning(f"PDF conversion error for {file_path.name}: {str(e)}")
        return False


def process_file(file_path: Path, input_dir: Path, output_dir: Path, pdf_output_dir: Path,
                alias_map: Dict, sorted_keys: List, compiled_patterns: Dict,
                logger: logging.Logger, remove_images: bool = True,
                generate_pdf: bool = True, timestamp_suffix: str = "") -> Dict:
    """
    Process a single file (anonymize and optionally convert to PDF)

    Returns:
        Dict with processing results
    """
    start_time = time.time()
    relative_path = file_path.relative_to(input_dir)

    # Determine file type
    extension = file_path.suffix.lower()

    # Handle legacy .doc and .ppt formats via LibreOffice conversion
    # Note: .xls files are handled directly with pandas (see routing below)
    if extension in ['.doc', '.ppt']:
        converted_path = convert_legacy_format(file_path, output_dir, logger)
        if converted_path is None:
            # Conversion failed - file will be copied as-is (not anonymized)
            return {
                'status': 'skipped',
                'replacements': 0,
                'images_removed': 0,
                'error': f"Legacy {extension} file cannot be converted - will be copied as-is (not anonymized)",
                'processing_time': time.time() - start_time
            }
        file_path = converted_path
        extension = file_path.suffix.lower()
        # Update relative path to reflect new extension
        relative_path = Path(str(relative_path).replace(relative_path.suffix, extension))

    # Determine output paths
    # Preserve entire folder structure including root folder name with optional timestamp
    root_folder_name = f"{input_dir.name}{timestamp_suffix}"
    output_path = output_dir / root_folder_name / relative_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    pdf_path = pdf_output_dir / root_folder_name / relative_path.with_suffix('.pdf')
    pdf_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        # Route to appropriate processor (v2.1: with detailed tracking)
        replacement_details = {}  # Will store per-document replacement details

        if extension == '.docx':
            logger.info(f"Processing DOCX: {relative_path}")
            replacements, images_removed, replacement_details = process_single_docx(
                str(file_path), str(output_path),
                alias_map, sorted_keys, logger,
                remove_images=remove_images,
                clear_headers_footers_flag=False,
                track_details=True
            )

        elif extension == '.pptx':
            logger.info(f"Processing PPTX: {relative_path}")
            replacements, images_removed, replacement_details = process_single_pptx(
                str(file_path), str(output_path),
                alias_map, sorted_keys, compiled_patterns, logger,
                remove_images=remove_images,
                track_details=True
            )

        elif extension == '.xlsx' or extension == '.xlsm':
            logger.info(f"Processing {extension.upper()}: {relative_path}")
            replacements, images_removed, replacement_details = process_single_xlsx(
                str(file_path), str(output_path),
                alias_map, sorted_keys, compiled_patterns, logger,
                remove_images=False,  # Excel doesn't support image removal
                track_details=True
            )

        elif extension == '.xls':
            logger.info(f"Processing XLS (legacy): {relative_path}")
            # For .xls files, output as .xlsx (converted format)
            output_path_xlsx = output_path.with_suffix('.xlsx')
            replacements, images_removed, replacement_details = process_single_xls(
                str(file_path), str(output_path_xlsx),
                alias_map, sorted_keys, compiled_patterns, logger,
                remove_images=False,  # Excel doesn't support image removal
                track_details=True
            )

        else:
            return {
                'status': 'skipped',
                'replacements': 0,
                'images_removed': 0,
                'error': f"Unsupported file type: {extension}",
                'processing_time': time.time() - start_time
            }

        # PDF conversion (optional)
        result_dict = {
            'status': 'success',
            'replacements': replacements,
            'images_removed': images_removed,
            'error': '',
            'processing_time': 0,  # Will be set below
            'replacement_details': replacement_details  # v2.1
        }

        # Only attempt PDF conversion and track result if PDF generation is enabled
        if generate_pdf and output_path.exists():
            pdf_success = convert_to_pdf(output_path, pdf_path.parent, logger)
            result_dict['pdf_success'] = pdf_success

        processing_time = time.time() - start_time
        result_dict['processing_time'] = processing_time
        logger.info(f"Completed {relative_path}: {replacements} replacements, "
                   f"{images_removed} images removed in {processing_time:.1f}s")

        return result_dict

    except Exception as e:
        error_msg = f"Processing error: {str(e)}"
        logger.error(f"Failed to process {relative_path}: {error_msg}")
        return {
            'status': 'failed',
            'replacements': 0,
            'images_removed': 0,
            'error': error_msg,
            'processing_time': time.time() - start_time
        }


def copy_non_processable_files(input_dir: Path, output_dir: Path, timestamp_suffix: str, logger: logging.Logger, stats: 'BatchStats') -> int:
    """
    Copy non-processable files (PDFs, images, etc.) to output directory
    to preserve complete folder structure.

    Args:
        stats: BatchStats object to track copied files

    Returns:
        Number of files copied
    """
    logger.info("Copying non-processable files to output...")

    # Extensions that should be copied as-is
    copy_extensions = {'.pdf', '.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.txt',
                      '.csv', '.json', '.xml', '.html', '.htm', '.zip', '.tar', '.gz'}

    # Tracker files to skip
    skip_patterns = ['tracker', 'anon tracker']

    files_copied = 0
    root_folder_name = f"{input_dir.name}{timestamp_suffix}"

    # Find all files in input directory
    for file_path in input_dir.rglob('*'):
        if not file_path.is_file():
            continue

        # Skip tracker files
        if any(pattern in file_path.name.lower() for pattern in skip_patterns):
            continue

        # Check if file should be copied
        if file_path.suffix.lower() in copy_extensions:
            # Calculate relative path and output path
            relative_path = file_path.relative_to(input_dir)
            output_path = output_dir / root_folder_name / relative_path

            # Create parent directories if needed
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Copy file
            try:
                shutil.copy2(file_path, output_path)
                files_copied += 1
                logger.debug(f"Copied: {relative_path}")

                # Track copied file
                stats.copied_files.append({
                    'original_path': str(relative_path),
                    'output_path': str(output_dir / root_folder_name / relative_path),
                    'file_type': file_path.suffix.lower(),
                    'filename': file_path.name
                })
            except Exception as e:
                logger.warning(f"Failed to copy {relative_path}: {str(e)}")

    logger.info(f"Copied {files_copied} non-processable files")
    return files_copied


def preserve_empty_folders(input_dir: Path, output_dir: Path, timestamp_suffix: str, logger: logging.Logger):
    """
    Create empty directories in output to match input structure.
    """
    logger.info("Preserving empty folder structure...")

    root_folder_name = f"{input_dir.name}{timestamp_suffix}"
    folders_created = 0

    # Find all directories in input
    for dir_path in input_dir.rglob('*'):
        if not dir_path.is_dir():
            continue

        # Calculate relative path and output path
        try:
            relative_path = dir_path.relative_to(input_dir)
            output_path = output_dir / root_folder_name / relative_path

            # Create directory if it doesn't exist
            if not output_path.exists():
                output_path.mkdir(parents=True, exist_ok=True)
                folders_created += 1
                logger.debug(f"Created empty folder: {relative_path}")
        except Exception as e:
            logger.warning(f"Failed to create directory {dir_path}: {str(e)}")

    logger.info(f"Created {folders_created} empty folders")


def discover_files(input_dir: Path, logger: logging.Logger) -> Dict[Path, List[Path]]:
    """
    Discover all processable files organized by top-level folder

    Returns:
        Dict mapping folder_path -> list of files in that folder tree
    """
    logger.info("Discovering files in folder structure...")

    # Supported extensions
    extensions = ['*.docx', '*.xlsx', '*.xlsm', '*.pptx', '*.doc', '*.xls', '*.ppt']

    # Find all files
    all_files = []
    for ext in extensions:
        all_files.extend(input_dir.rglob(ext))

    # Filter out tracker files
    all_files = [f for f in all_files if 'tracker' not in f.name.lower() and 'anon tracker' not in f.name.lower()]

    # Organize by top-level folder
    folder_files = defaultdict(list)

    for file_path in all_files:
        try:
            relative = file_path.relative_to(input_dir)

            # Get top-level folder (or root if file is directly in input_dir)
            if len(relative.parts) > 1:
                top_folder = input_dir / relative.parts[0]
            else:
                top_folder = input_dir

            folder_files[top_folder].append(file_path)
        except ValueError:
            logger.warning(f"Skipping file outside input directory: {file_path}")

    # Log discovery results
    logger.info(f"Discovered {len(all_files)} processable files in {len(folder_files)} top-level folders")
    for folder, files in sorted(folder_files.items()):
        try:
            folder_rel = folder.relative_to(input_dir)
        except ValueError:
            folder_rel = folder
        logger.info(f"  {folder_rel}: {len(files)} files")

    return folder_files


def generate_excel_report(stats: BatchStats, report_path: Path, alias_map: Dict, logger: logging.Logger):
    """Generate comprehensive Excel report with replacement and copied files details"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        logger.info(f"Generating Excel report: {report_path}")

        wb = Workbook()

        # Sheet 1: File Details
        ws_files = wb.active
        ws_files.title = "File Details"

        headers = ['File Path', 'Folder', 'Filename', 'Type', 'Status',
                  'Replacements', 'Images Removed', 'Processing Time (s)', 'Error']
        ws_files.append(headers)

        # Style headers
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws_files.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Add file data
        for detail in stats.file_details:
            ws_files.append([
                detail['file_path'],
                detail['folder'],
                detail['filename'],
                detail['extension'],
                detail['status'],
                detail['replacements'],
                detail['images_removed'],
                round(detail['processing_time'], 2),
                detail['error']
            ])

        # Auto-size columns
        for col_num in range(1, len(headers) + 1):
            ws_files.column_dimensions[get_column_letter(col_num)].width = 15
        ws_files.column_dimensions['A'].width = 50
        ws_files.column_dimensions['I'].width = 40

        # Sheet 2: Folder Summary
        ws_folders = wb.create_sheet("Folder Summary")

        folder_headers = ['Folder', 'Total Files', 'Succeeded', 'Failed', 'Skipped',
                         'Replacements', 'Images Removed', 'Success Rate (%)']
        ws_folders.append(folder_headers)

        # Style headers
        for col_num, header in enumerate(folder_headers, 1):
            cell = ws_folders.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Add folder data
        for folder_name, folder_data in sorted(stats.folder_stats.items()):
            success_rate = (folder_data['succeeded'] / max(folder_data['files'], 1)) * 100
            ws_folders.append([
                folder_name,
                folder_data['files'],
                folder_data['succeeded'],
                folder_data['failed'],
                folder_data['skipped'],
                folder_data['replacements'],
                folder_data['images_removed'],
                round(success_rate, 1)
            ])

        # Auto-size columns
        for col_num in range(1, len(folder_headers) + 1):
            ws_folders.column_dimensions[get_column_letter(col_num)].width = 15
        ws_folders.column_dimensions['A'].width = 40

        # Sheet 3: Run Summary
        ws_summary = wb.create_sheet("Run Summary")

        summary_data = [
            ['Metric', 'Value'],
            ['Total Files Processed', stats.files_processed],
            ['Files Succeeded', stats.files_succeeded],
            ['Files Failed', stats.files_failed],
            ['Files Skipped', stats.files_skipped],
            ['Success Rate (%)', round((stats.files_succeeded / max(stats.files_processed, 1)) * 100, 1)],
            ['', ''],
            ['Total Replacements', stats.total_replacements],
            ['Total Images Removed', stats.total_images_removed],
            ['', ''],
            ['PDF Conversions Succeeded', stats.pdf_successes],
            ['PDF Conversions Failed', stats.pdf_failures],
            ['', ''],
            ['Processing Started', stats.start_time.strftime('%Y-%m-%d %H:%M:%S')],
            ['Processing Completed', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Processing Time', stats.get_elapsed_time()]
        ]

        for row in summary_data:
            ws_summary.append(row)

        # Style summary
        for row_num in range(1, len(summary_data) + 1):
            cell_a = ws_summary.cell(row_num, 1)
            cell_a.font = Font(bold=True)
            cell_a.alignment = Alignment(horizontal='left')

        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 30

        # Sheet 4: Error Log (if any errors)
        if stats.error_log:
            ws_errors = wb.create_sheet("Error Log")
            error_headers = ['Timestamp', 'File Path', 'Error Message']
            ws_errors.append(error_headers)

            # Style headers
            for col_num, header in enumerate(error_headers, 1):
                cell = ws_errors.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add error data
            for error in stats.error_log:
                ws_errors.append([
                    error['timestamp'],
                    error['file_path'],
                    error['error']
                ])

            ws_errors.column_dimensions['A'].width = 20
            ws_errors.column_dimensions['B'].width = 50
            ws_errors.column_dimensions['C'].width = 60

        # Sheet 5: Anonymization Mappings
        ws_mappings = wb.create_sheet("Anonymization Mappings")
        mapping_headers = ['Original Text', 'Replacement Text', 'Action Type']
        ws_mappings.append(mapping_headers)

        # Style headers
        for col_num, header in enumerate(mapping_headers, 1):
            cell = ws_mappings.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Add mapping data
        for original, replacement in sorted(alias_map.items()):
            action_type = "DELETION (blank)" if replacement == "" else "REPLACEMENT"
            ws_mappings.append([original, replacement if replacement != "" else "[DELETED]", action_type])

        ws_mappings.column_dimensions['A'].width = 40
        ws_mappings.column_dimensions['B'].width = 40
        ws_mappings.column_dimensions['C'].width = 20

        # Sheet 6: Copied Files (Non-Processable)
        if stats.copied_files:
            ws_copied = wb.create_sheet("Copied Files")
            copied_headers = ['Original Path', 'Output Location', 'File Type', 'Filename']
            ws_copied.append(copied_headers)

            # Style headers
            for col_num, header in enumerate(copied_headers, 1):
                cell = ws_copied.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add copied file data
            for file_info in stats.copied_files:
                ws_copied.append([
                    file_info['original_path'],
                    file_info['output_path'],
                    file_info['file_type'],
                    file_info['filename']
                ])

            ws_copied.column_dimensions['A'].width = 50
            ws_copied.column_dimensions['B'].width = 60
            ws_copied.column_dimensions['C'].width = 15
            ws_copied.column_dimensions['D'].width = 40

        # Sheet 7: Detailed Replacements by Document (v2.1)
        if stats.file_replacement_details:
            ws_details = wb.create_sheet("Detailed Replacements")
            details_headers = ["File Path", "Directory", "Document", "Original Text", "Replacement", "Occurrences", "Action Type"]
            ws_details.append(details_headers)

            # Style headers
            for col_num, header in enumerate(details_headers, 1):
                cell = ws_details.cell(1, col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add detailed replacement data
            rows = []
            for file_info in stats.file_replacement_details:
                file_path = file_info['file_path']
                directory = file_info['directory']
                filename = file_info['filename']
                details = file_info['details']  # {original: count, ...}

                # Create a row for each original → replacement
                for original, count in sorted(details.items(), key=lambda x: x[1], reverse=True):
                    # Look up replacement from alias_map
                    replacement = alias_map.get(original, "[UNKNOWN]")
                    action_type = "DELETION (blank)" if replacement == "" else "REPLACEMENT"
                    display_replacement = "[DELETED]" if replacement == "" else replacement

                    rows.append([
                        file_path,
                        directory,
                        filename,
                        original,
                        display_replacement,
                        count,
                        action_type
                    ])

            # Sort rows by file path, then by occurrences (descending)
            rows.sort(key=lambda x: (x[0], -x[5]))

            # Add all rows
            for row in rows:
                ws_details.append(row)

            # Set column widths
            ws_details.column_dimensions['A'].width = 50  # File Path
            ws_details.column_dimensions['B'].width = 25  # Directory
            ws_details.column_dimensions['C'].width = 40  # Document
            ws_details.column_dimensions['D'].width = 30  # Original Text
            ws_details.column_dimensions['E'].width = 30  # Replacement
            ws_details.column_dimensions['F'].width = 12  # Occurrences
            ws_details.column_dimensions['G'].width = 20  # Action Type

        # Save workbook
        wb.save(report_path)
        logger.info(f"Excel report generated successfully: {report_path}")

    except Exception as e:
        logger.error(f"Failed to generate Excel report: {str(e)}")


def main():
    """Main batch processing function"""
    parser = argparse.ArgumentParser(
        description='Batch anonymize entire folder structures',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Process with prompting for each folder
  python batch_anonymize.py --input ./input/Project\\ Nautilus --output ./output --tracker ./tracker/Anon\\ Tracker\\ -\\ Nautilus.xlsx

  # Dry run (preview only, no processing)
  python batch_anonymize.py --input ./input/Project\\ Nautilus --output ./output --tracker ./tracker/Anon\\ Tracker\\ -\\ Nautilus.xlsx --dry-run

  # Skip PDF generation (faster)
  python batch_anonymize.py --input ./input/Project\\ Nautilus --output ./output --tracker ./tracker/Anon\\ Tracker\\ -\\ Nautilus.xlsx --no-pdf
        """
    )

    parser.add_argument('--input', required=True, type=str,
                       help='Input directory containing files to process')
    parser.add_argument('--output', required=True, type=str,
                       help='Output directory for anonymized files')
    parser.add_argument('--tracker', required=True, type=str,
                       help='Path to Excel anonymization tracker file')
    parser.add_argument('--pdf-output', type=str, default=None,
                       help='Output directory for PDF files (default: ./pdf_output)')
    parser.add_argument('--no-pdf', action='store_true',
                       help='Skip PDF conversion (faster processing)')
    parser.add_argument('--dry-run', action='store_true',
                       help='Preview files to be processed without actually processing')
    parser.add_argument('--auto-yes-images', action='store_true',
                       help='Automatically remove images from all folders (no prompting)')
    parser.add_argument('--auto-no-images', action='store_true',
                       help='Automatically preserve images in all folders (no prompting)')
    parser.add_argument('--timestamp-output', action='store_true',
                       help='Add timestamp to output folder names (prevents overwriting previous runs)')

    args = parser.parse_args()

    # Setup paths
    input_dir = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    tracker_path = Path(args.tracker).resolve()

    # Setup PDF output directory
    if args.pdf_output:
        pdf_output_dir = Path(args.pdf_output).resolve()
    else:
        pdf_output_dir = Path(__file__).parent / 'pdf_output'

    # Generate timestamp for folder naming if requested
    timestamp_suffix = ""
    if args.timestamp_output:
        timestamp_suffix = f"_{datetime.now().strftime('%Y%m%d_%H%M%S')}"

    # Validate inputs
    if not input_dir.exists():
        print(f"{Colors.RED}Error: Input directory does not exist: {input_dir}{Colors.ENDC}")
        sys.exit(1)

    if not tracker_path.exists():
        print(f"{Colors.RED}Error: Tracker file does not exist: {tracker_path}{Colors.ENDC}")
        sys.exit(1)

    # Setup logging
    log_dir = Path(__file__).parent / 'logs'
    logger = setup_logging(log_dir)

    # Print banner
    print(f"""
{Colors.BOLD}{Colors.CYAN}╔════════════════════════════════════════════════════════════╗
║       FOLDER BATCH ANONYMIZATION SYSTEM v2.0           ║
║       Production-Grade Document Processing             ║
╚════════════════════════════════════════════════════════════╝{Colors.ENDC}

{Colors.BOLD}Configuration:{Colors.ENDC}
  Input:       {input_dir}
  Output:      {output_dir}
  Tracker:     {tracker_path}
  PDF Output:  {pdf_output_dir if not args.no_pdf else 'DISABLED'}
  Mode:        {'DRY RUN (Preview Only)' if args.dry_run else 'LIVE PROCESSING'}
""")

    # Load anonymization mappings
    print(f"{Colors.BOLD}Loading anonymization tracker...{Colors.ENDC}")
    try:
        alias_map = load_aliases_from_excel(str(tracker_path))
        sorted_keys = categorize_and_sort_aliases(alias_map)
        compiled_patterns = precompile_patterns(alias_map)
        print(f"{Colors.GREEN}✓ Loaded {len(alias_map)} anonymization mappings{Colors.ENDC}")
        logger.info(f"Loaded {len(alias_map)} anonymization mappings from {tracker_path}")
    except Exception as e:
        print(f"{Colors.RED}✗ Failed to load tracker: {str(e)}{Colors.ENDC}")
        logger.error(f"Failed to load tracker: {str(e)}")
        sys.exit(1)

    # Discover files
    print(f"\n{Colors.BOLD}Discovering files...{Colors.ENDC}")
    folder_files = discover_files(input_dir, logger)

    if not folder_files:
        print(f"{Colors.YELLOW}No processable files found in {input_dir}{Colors.ENDC}")
        sys.exit(0)

    total_files = sum(len(files) for files in folder_files.values())
    print(f"{Colors.GREEN}✓ Found {total_files} files in {len(folder_files)} top-level folders{Colors.ENDC}")

    # Dry run mode
    if args.dry_run:
        print(f"\n{Colors.YELLOW}{Colors.BOLD}DRY RUN MODE - Preview Only{Colors.ENDC}")
        total_warnings = 0
        for folder_path in sorted(folder_files.keys()):
            folder_info = get_folder_info(folder_path, input_dir)
            print(f"\n{Colors.BOLD}Folder:{Colors.ENDC} {folder_info['path']}")
            print(f"  Processable files: {folder_info['file_count']}")
            type_str = " | ".join([f"{k}: {v}" for k, v in folder_info['type_counts'].items() if v > 0])
            if type_str:
                print(f"  Types: {type_str}")

            # Show info for non-processable files
            if folder_info['has_warnings']:
                non_proc_str = " | ".join([f"{k}: {v}" for k, v in folder_info['non_processable_counts'].items()])
                print(f"  {Colors.CYAN}ℹ Non-processable: {non_proc_str} (will be COPIED as-is){Colors.ENDC}")
                total_warnings += folder_info['non_processable_total']

        print(f"\n{Colors.CYAN}Total files that would be processed: {total_files}{Colors.ENDC}")
        if total_warnings > 0:
            print(f"{Colors.CYAN}Total non-processable files that would be copied as-is: {total_warnings}{Colors.ENDC}")
        print(f"{Colors.YELLOW}Run without --dry-run to process files{Colors.ENDC}")
        sys.exit(0)

    # Initialize stats and progress
    stats = BatchStats()

    # Determine auto-mode for image removal
    auto_mode = None
    if args.auto_yes_images:
        auto_mode = True
        print(f"{Colors.CYAN}Auto-remove images mode enabled{Colors.ENDC}")
    elif args.auto_no_images:
        auto_mode = False
        print(f"{Colors.CYAN}Auto-preserve images mode enabled{Colors.ENDC}")

    # Process each folder
    print(f"\n{Colors.BOLD}{Colors.GREEN}Starting batch processing...{Colors.ENDC}\n")

    progress = ProgressDisplay(total_files)

    for folder_path in sorted(folder_files.keys()):
        files = folder_files[folder_path]

        # Get folder info and prompt for image removal
        folder_info = get_folder_info(folder_path, input_dir)
        remove_images, auto_mode = prompt_for_image_removal(folder_info, auto_mode)

        # Check if user chose to skip
        if remove_images is None:
            logger.info(f"Skipping folder: {folder_info['path']}")
            for file_path in files:
                relative_path = file_path.relative_to(input_dir)
                stats.add_file_result(file_path, relative_path, 'skipped',
                                    error_msg="User skipped folder")
            continue

        # Process all files in this folder
        for file_path in files:
            relative_path = file_path.relative_to(input_dir)

            # Process file
            result = process_file(
                file_path, input_dir, output_dir, pdf_output_dir,
                alias_map, sorted_keys, compiled_patterns,
                logger, remove_images=remove_images,
                generate_pdf=not args.no_pdf,
                timestamp_suffix=timestamp_suffix
            )

            # Update stats
            stats.add_file_result(
                file_path, relative_path, result['status'],
                result['replacements'], result['images_removed'],
                result['processing_time'], result.get('error', ''),
                replacement_details=result.get('replacement_details', {})
            )

            # Track PDF conversion
            if not args.no_pdf and 'pdf_success' in result:
                stats.add_pdf_result(result['pdf_success'])

            # Update progress display
            progress.update(str(relative_path), stats)

    # Copy non-processable files and preserve empty folders
    print(f"\n{Colors.BOLD}Finalizing folder structure...{Colors.ENDC}")
    files_copied = copy_non_processable_files(input_dir, output_dir, timestamp_suffix, logger, stats)
    preserve_empty_folders(input_dir, output_dir, timestamp_suffix, logger)
    print(f"{Colors.GREEN}✓ Copied {files_copied} non-processable files and preserved folder structure{Colors.ENDC}")

    # Print final summary
    print(f"\n\n{stats.get_summary(include_pdf=not args.no_pdf)}")

    # Generate Excel report
    report_dir = Path(__file__).parent / 'reports'
    report_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_path = report_dir / f"batch_report_{timestamp}.xlsx"

    print(f"\n{Colors.BOLD}Generating comprehensive report...{Colors.ENDC}")
    generate_excel_report(stats, report_path, alias_map, logger)
    print(f"{Colors.GREEN}✓ Report saved: {report_path}{Colors.ENDC}")

    # Cleanup temp conversion files
    temp_dir = output_dir / '.temp_conversions'
    if temp_dir.exists():
        shutil.rmtree(temp_dir)
        logger.info("Cleaned up temporary conversion files")

    logger.info("Batch processing completed successfully")
    print(f"\n{Colors.BOLD}{Colors.GREEN}✓ Batch processing completed!{Colors.ENDC}\n")


if __name__ == '__main__':
    main()
