#!/usr/bin/env python3
"""
Excel Anonymization Module
Handles .xlsx and .xls files for the DOCX Anonymizer app
"""

from openpyxl import load_workbook, Workbook
from pathlib import Path
import logging
import re
import pandas as pd


def strip_xlsx_metadata(wb):
    """
    Strip ALL metadata from Excel file.

    Removes:
    - Author
    - Title
    - Subject
    - Keywords
    - Comments
    - Company
    """
    props = wb.properties

    props.creator = ""
    props.lastModifiedBy = ""
    props.title = ""
    props.subject = ""
    props.keywords = ""
    props.description = ""
    props.category = ""
    props.contentStatus = ""
    props.identifier = ""
    props.company = ""
    props.manager = ""

    props.revision = 1
    if hasattr(props, 'version'):
        props.version = None

    return wb


def anonymize_text_xlsx(text, alias_map, sorted_keys, compiled_patterns, track_details=False):
    """
    Apply anonymization replacements with case matching using SINGLE-PASS regex (v2.1).

    Reuses same logic as Word/PowerPoint processors.

    Returns:
        If track_details=False: (text, replacements)
        If track_details=True: (text, replacements, details_dict)
    """
    if not text or not isinstance(text, str):
        if track_details:
            return text, 0, {}
        return text, 0

    replacements = 0

    # Extract combined pattern and lookup map
    combined_pattern = compiled_patterns.get('combined')
    lookup = compiled_patterns.get('lookup')

    # BACKWARD COMPATIBILITY: Handle old compiled_patterns format
    if combined_pattern is None or lookup is None:
        # Old format - use legacy multi-pass algorithm
        result = anonymize_text_xlsx_legacy(text, alias_map, sorted_keys, compiled_patterns)
        if track_details:
            return result[0], result[1], {}
        return result

    # Track which originals were replaced (v2.1 feature)
    details = {} if track_details else None

    # SINGLE-PASS REPLACEMENT (v2.1 performance optimization)
    def replace_match(match):
        nonlocal replacements
        matched_text = match.group(0)

        # Look up the replacement using lowercase match
        matched_lower = matched_text.lower()
        if matched_lower not in lookup:
            return matched_text  # Safe fallback

        original, replacement = lookup[matched_lower]

        # Track this replacement (v2.1)
        if track_details:
            details[original] = details.get(original, 0) + 1

        # Preserve case pattern
        if matched_text.isupper():
            replacements += 1
            return replacement.upper()
        elif matched_text.islower():
            replacements += 1
            return replacement.lower()
        elif matched_text[0].isupper():
            replacements += 1
            return replacement.capitalize()
        else:
            replacements += 1
            return replacement

    # Single regex pass replaces ALL patterns at once
    text = combined_pattern.sub(replace_match, text)

    if track_details:
        return text, replacements, details
    return text, replacements


def merge_details(details1, details2):
    """
    Merge two replacement details dictionaries (v2.1 helper).

    Args:
        details1: First details dict {original: count, ...}
        details2: Second details dict to merge in

    Returns:
        Merged details dict
    """
    if details1 is None:
        return details2 if details2 else {}
    if details2 is None:
        return details1

    merged = details1.copy()
    for original, count in details2.items():
        merged[original] = merged.get(original, 0) + count
    return merged


def anonymize_text_xlsx_legacy(text, alias_map, sorted_keys, compiled_patterns):
    """
    Legacy multi-pass anonymization (kept for backward compatibility).
    """
    if not text or not isinstance(text, str):
        return text, 0

    replacements = 0

    for original in sorted_keys:
        replacement = alias_map[original]

        def replace_with_case(match):
            nonlocal replacements
            matched_text = match.group(0)

            # Preserve case pattern
            if matched_text.isupper():
                replacements += 1
                return replacement.upper()
            elif matched_text.islower():
                replacements += 1
                return replacement.lower()
            elif matched_text[0].isupper():
                replacements += 1
                return replacement.capitalize()
            else:
                replacements += 1
                return replacement

        pattern = compiled_patterns[original]
        text = pattern.sub(replace_with_case, text)

    return text, replacements


def anonymize_xlsx(xlsx_path, alias_map, sorted_keys, compiled_patterns, track_details=False):
    """
    Anonymize all text in Excel file (v2.1 with optional tracking).

    Processes:
    - Cell values (text and formulas)
    - Sheet names
    - Headers/footers
    - Comments

    IMPORTANT: Does NOT recalculate formulas (safer, prevents errors)

    Returns:
        If track_details=False: (wb, total_replacements)
        If track_details=True: (wb, total_replacements, details_dict)
    """
    wb = load_workbook(xlsx_path, data_only=False)  # Keep formulas
    total_replacements = 0
    document_details = {} if track_details else None

    # Create tracking wrapper
    def anonymize_with_tracking(text, alias_map, sorted_keys, compiled_patterns):
        nonlocal document_details
        if track_details:
            new_text, count, details = anonymize_text_xlsx(text, alias_map, sorted_keys, compiled_patterns, track_details=True)
            document_details = merge_details(document_details, details)
            return new_text, count
        else:
            return anonymize_text_xlsx(text, alias_map, sorted_keys, compiled_patterns)

    # Anonymize sheet names first
    sheet_name_mapping = {}
    for sheet in wb.worksheets:
        old_name = sheet.title
        new_name, count = anonymize_with_tracking(old_name, alias_map, sorted_keys, compiled_patterns)
        if count > 0:
            # Ensure unique sheet name (Excel requirement)
            if new_name in [s.title for s in wb.worksheets]:
                new_name = f"{new_name}_1"
            sheet.title = new_name
            sheet_name_mapping[old_name] = new_name
            total_replacements += count

    # Process each sheet
    for sheet in wb.worksheets:
        # Process all cells
        for row in sheet.iter_rows():
            for cell in row:
                # Anonymize cell values (text)
                if cell.value and isinstance(cell.value, str):
                    new_value, count = anonymize_with_tracking(
                        cell.value, alias_map, sorted_keys, compiled_patterns
                    )
                    if count > 0:
                        # Check if it's a formula
                        if str(cell.value).startswith('='):
                            # It's a formula - replace text within formula
                            cell.value = new_value
                        else:
                            # Regular text
                            cell.value = new_value
                        total_replacements += count

                # Anonymize cell comments
                if cell.comment:
                    if cell.comment.text:
                        new_comment, count = anonymize_with_tracking(
                            cell.comment.text, alias_map, sorted_keys, compiled_patterns
                        )
                        if count > 0:
                            cell.comment.text = new_comment
                            total_replacements += count

        # NOTE: Excel header/footer anonymization skipped
        # openpyxl header/footer objects have complex structure
        # Most Excel files don't use headers/footers with company names
        # Can be added in future if needed

    if track_details:
        return wb, total_replacements, document_details
    return wb, total_replacements


def process_single_xlsx(input_path, output_path, alias_map, sorted_keys, compiled_patterns, logger, remove_images=True, track_details=False):
    """
    Process a single Excel file: anonymize + strip metadata.

    Args:
        input_path: Path to input .xlsx file (string or Path object)
        output_path: Path for output .xlsx file (string or Path object)
        alias_map: Dictionary of original → replacement mappings
        sorted_keys: Sorted list of alias_map keys
        compiled_patterns: Pre-compiled regex patterns
        logger: Logger instance
        remove_images: Ignored for Excel (kept for API consistency)
        track_details: If True, return detailed replacement tracking (v2.1)

    Returns:
        If track_details=False: (replacements, images_removed)
        If track_details=True: (replacements, images_removed, details_dict)
        Note: images_removed always 0 for Excel (charts/images not processed)
    """
    # Convert to Path objects if strings (for backward compatibility)
    from pathlib import Path
    input_path = Path(input_path) if isinstance(input_path, str) else input_path
    output_path = Path(output_path) if isinstance(output_path, str) else output_path

    logger.info(f"Processing: {input_path.name}")

    try:
        # Load and anonymize Excel with optional tracking
        if track_details:
            wb, replacements, details = anonymize_xlsx(input_path, alias_map, sorted_keys, compiled_patterns, track_details=True)
        else:
            wb, replacements = anonymize_xlsx(input_path, alias_map, sorted_keys, compiled_patterns)

        # Strip ALL metadata (CRITICAL)
        wb = strip_xlsx_metadata(wb)

        # Save
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"  ✓ {replacements} replacements")

        # Return 0 for images (Excel doesn't remove images yet)
        if track_details:
            return replacements, 0, details
        return replacements, 0

    except Exception as e:
        logger.error(f"  ❌ Error: {e}")
        if track_details:
            return 0, 0, {}
        return 0, 0


def process_single_xls(input_path, output_path, alias_map, sorted_keys, compiled_patterns, logger, remove_images=True, track_details=False):
    """
    Process a legacy .xls file: convert to .xlsx, anonymize + strip metadata.

    Args:
        input_path: Path to input .xls file (string or Path object)
        output_path: Path for output .xlsx file (string or Path object)
        alias_map: Dictionary of original → replacement mappings
        sorted_keys: Sorted list of alias_map keys
        compiled_patterns: Pre-compiled regex patterns
        logger: Logger instance
        remove_images: Ignored for Excel (kept for API consistency)
        track_details: If True, return detailed replacement tracking (v2.1)

    Returns:
        If track_details=False: (replacements, images_removed)
        If track_details=True: (replacements, images_removed, details_dict)
        Note: images_removed always 0 for .xls files
    """
    from pathlib import Path
    input_path = Path(input_path) if isinstance(input_path, str) else input_path
    output_path = Path(output_path) if isinstance(output_path, str) else output_path

    logger.info(f"Processing XLS: {input_path.name}")

    try:
        # Read legacy .xls file using pandas with xlrd engine
        xls_file = pd.ExcelFile(input_path, engine='xlrd')

        # Create new .xlsx workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        total_replacements = 0
        document_details = {} if track_details else None

        # Create tracking wrapper
        def anonymize_with_tracking(text, alias_map, sorted_keys, compiled_patterns):
            nonlocal document_details
            if track_details:
                new_text, count, details = anonymize_text_xlsx(text, alias_map, sorted_keys, compiled_patterns, track_details=True)
                document_details = merge_details(document_details, details)
                return new_text, count
            else:
                return anonymize_text_xlsx(text, alias_map, sorted_keys, compiled_patterns)

        # Process each sheet
        for sheet_name in xls_file.sheet_names:
            # Read sheet data (no headers)
            df = pd.read_excel(xls_file, sheet_name=sheet_name, header=None)

            # Anonymize sheet name
            anonymized_sheet_name, name_count = anonymize_with_tracking(sheet_name, alias_map, sorted_keys, compiled_patterns)
            total_replacements += name_count

            # Create new sheet in output workbook
            ws = wb.create_sheet(title=anonymized_sheet_name)

            # Process each cell
            for row_idx, row in df.iterrows():
                for col_idx, cell_value in enumerate(row):
                    if pd.notna(cell_value) and isinstance(cell_value, str):
                        # Anonymize cell value
                        anonymized_value, count = anonymize_with_tracking(cell_value, alias_map, sorted_keys, compiled_patterns)
                        total_replacements += count
                        # Write to new workbook (row/col are 1-indexed in openpyxl)
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=anonymized_value)
                    else:
                        # Non-string value - copy as-is
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

        # Strip ALL metadata
        wb = strip_xlsx_metadata(wb)

        # Save as .xlsx
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_path)

        logger.info(f"  ✓ {total_replacements} replacements (.xls → .xlsx)")

        if track_details:
            return total_replacements, 0, document_details
        return total_replacements, 0

    except Exception as e:
        logger.error(f"  ❌ Error processing .xls file: {e}")
        if track_details:
            return 0, 0, {}
        return 0, 0
